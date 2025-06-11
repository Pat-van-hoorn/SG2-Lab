import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
import random
import openpyxl

'''------INTITALISE--SIMULATION--PARAMETERS-----------'''


# Vessel Geometry Parameters (SI units)
R = 0.03      # Radius of cylinder (m)
L = 0.06;       # Height of cylinder (m)
fill_fraction = 0.5   #Fraction of total volume filled with water
    
    #Calculate vessel volumes
V_vessel = np.pi * R**2 * L             # Total vessel volume (m^3)
V_water = fill_fraction * V_vessel   # Water volume (m^3)
    
    # Water Properties
rho_water = 1000    # Water density (kg/m^3)
c_water = 4186      # Specific heat capacity (J/(kg*K))
m_water = rho_water * V_water# Mass of water    
C_water = m_water * c_water         # Water heat capacity (J/K)
    
    # Vessel (Glass) Properties
mass_vessel = 0.05     # Mass of the glass vessel (kg)
c_vessel = 840         # Specific heat capacity of glass (J/(kg*K))
C_glass = mass_vessel * c_vessel   # Glass heat capacity (J/K)
    
    # Area Calculations
    # Cylindrical (side) surface (where the heating mat is wrapped)
A_side_total = 2 * np.pi * R * L;
    
    # Water contacts the vessel on the bottom and the side up to the water level.
h_water = fill_fraction * L  # height of the water column (m)
A_side_water = 2 * np.pi * R * h_water # portion of side in contact with water
A_bottom = np.pi * R**2              # bottom area (always water contact)
A_water_contact = A_side_water + A_bottom  # Total area where glass contacts water
    
    # The environmental losses occur from the vessel surfaces not in contact with water.
    # Side area not in
A_side_env = A_side_total - A_side_water
    # Top area: if the vessel is not full, the top is exposed to air; if full, set to zero.
if fill_fraction < 1:
    A_top = np.pi * R**2
else:
    A_top = 0

    # Bottom area: 
A_bottom = np.pi * R**2
    
A_env = A_side_env + A_top + A_bottom  # Total area exposed to ambient
    
    # Heat Transfer Parameters
    # Heater: a fixed power is applied when ON.
Q_heater_const = 3.5  #Heater power (W) when on (Orignal = 10)
    
    # Heat transfer coefficients (W/(m^2*K))
h_gw = 25   # from glass to water (Original = 50)
h_loss = 3.8  # from glass to ambient (Original = 10)
    
T_ambient = 22   # Ambient temperature (°C)
T_initial = 28.5

#Bacterial Growth Parameters

doubling_time = 46.5 # Doubling time in minutes


alpha_max = np.log(2)/(doubling_time*60) #Maximum Growth Rate

B = 2*10**6 #Maximum Population Size

A_initial = 2000 #Initial Popuation, Usually 2000

# A_initial = 25000 #Comment out if not looking at chemostat

C = 4.7*10**-7 #Proporionality Constant between Popluation size and OD Readings

#Noise Parameters

noise = True

if noise:
    T_glass_sigma = 0.08
    T_measurement_sigma = 0.03
    Growth_Rate_sigma_mu = 1 #Models intrinsic noise in gene expression - always set to 1!!
    
    OD_measurement_sigma = 0.001

else:
    T_glass_sigma, T_measurement_sigma, Growth_Rate_sigma_mu, OD_measurement_sigma = 0,0,0,0


# Hysteresis band (°C) to avoid rapid switching
Hysteresis_band = 0.5
Target_temp = 30
Control_delay = 5 #For bang-bang control, seconds

def temperature_control(T_water,heater_power,k, dt, control_type = 'Bang-Bang'):

    if control_type == 'Hysterisis':
    
        #Hysterisis Control
    
        if heater_power[k-1] == 1.0:
            if T_water[k] > Target_temp + Hysteresis_band:
                return 0.0
            else:
                return 1.0
        else:
            if T_water[k] < Target_temp - Hysteresis_band:
                return 1.0
            else:
                return 0.0
    if control_type == 'Bang-Bang':
        if heater_power[k-1] == 1.0 and np.sum(T_water[int(k-1-Control_delay*dt)]) > Target_temp:
            return 0.0
        elif heater_power[k-1] == 0.0 and T_water[int(k-1-Control_delay*dt)] < Target_temp:
            return 1.0
        else:
            return heater_power[k-1]


def OD_control(OD_readings,k, dt, control_type):

    if control_type == 'None':
        return 0

    #Open_Loop control
    #Set both motor powers to 65, corresponds to fluid inflow/outflow rate of 0.02ml/s

    if control_type == 'Open-Loop':
        in_out_flowrate = 0.02*10**-6  #Measure in m^3/s

        

    return in_out_flowrate
    
    
                        
def plot_temp_control():
    Times, T_water, measured_T,heater_state, OD, A = run_sim(1, 3600)
    
    fig, ax1 = plt.subplots()
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel("Temperatures (°C)", color = 'red')
    ax1.plot(Times, T_water, label = 'Temperatures', color = 'red')
    #plt.ylim(25,33)
    ax1.legend()

    ax2 = ax1.twinx()
    
    ax2.plot(Times, heater_state, label = 'Heater State', color = 'blue')
    ax2.set_ylabel('Heater State', color = 'blue')
    plt.ylim(-1,2)
    plt.title("Heater State & Temperature")
    ax2.legend('Heater State', loc="lower right")
    plt.grid()
    plt.show()    

def plot_temp_OD():
    Times, T_water, measured_T,heater_state, OD, A = run_sim(1, 30000)
    fig, ax1 = plt.subplots()
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel("Temperatures (°C)", color = 'red')
    ax1.plot(Times, T_water, label = 'Temperatures', color = 'red')
    plt.ylim(20,35)
    ax1.legend()

    ax2 = ax1.twinx()
    
    ax2.plot(Times, A, label = 'OD Readings', color = 'blue')
    ax2.set_ylabel('Bacteria Population', color = 'blue')
    plt.title("OD_Readings & Temperature")
    ax2.legend('OD_Readings', loc="lower right")
    plt.grid()
    plt.show()

def get_data(file_name, InterpolateOD=False):

    #Get Data from excel spreadsheet
    
    spread = openpyxl.load_workbook(file_name)
    ws = spread.active
    m_ODReadings = np.array([ws.cell(row=i,column=4).value for i in range(1,ws.max_row+1)])
    m_Temperatures = np.array([ws.cell(row=i,column=5).value for i in range(1,ws.max_row+1)])
    m_Times = np.array([ws.cell(row=i,column=8).value for i in range(1,ws.max_row+1)])
    m_RelayState = np.array([ws.cell(row=i,column=6).value for i in range(1,ws.max_row+1)])


    #When the relay is on, our OD Arificially increases. We can interpolate our data
    #within these intervals in order to remove this effect.
    
    if InterpolateOD == True:
        for i in range(int(np.size(m_Times))):
            if m_RelayState[i] == 1 and m_RelayState[i-1] == 0:
                j = 0
                while m_RelayState[i+j]==1:
                    j+=1
                if i+j+8<int(np.size(m_Times)):
                    grad = (m_ODReadings[i+j+8]-m_ODReadings[i-1])/(j+9)
                    for x in range(j+9):
                        m_ODReadings[i+x-1] = m_ODReadings[i-1]+x*grad

    return m_ODReadings, m_Temperatures, m_Times, m_RelayState

def compare_plots_temp(plot_heater_state=False):
    m_ODReadings, m_Temperatures, m_Times, m_heater_state = get_data('Growth_Curve_Stats.xlsx')
    
    Times, T_water, measured_T,heater_state, ODReadings, A = run_sim(1, int(np.size(m_Times)))
    
    fig, ax1 = plt.subplots()
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel("Temperatures (°C)",)
    
    ax1.plot(Times, measured_T, label = 'Simulated Temperature', color = 'red')
    ax1.plot(Times, m_Temperatures, label = 'Measured Temperature', color = 'purple')
    
    plt.ylim(28,33)
    ax1.legend()
    plt.xlim(20000,30000)
    #Optional: also plot the heater state\
    if plot_heater_state:
        ax2 = ax1.twinx()
        
        ax2.plot(Times, heater_state, label = 'Simulated Heater State', color = 'blue')
        ax2.plot(Times, m_heater_state, label = 'Measured Heater State', color = 'black')
        ax2.set_ylabel('Heater State',)
        plt.ylim(-1,5)
        plt.title("Heater State & Temperature over time")
        ax2.legend(loc='lower right')
    else:
        plt.title("Temperature over time")
    
    ax1.grid()
    plt.show()
    
def compare_plots_growth_curve():
    m_ODReadings, m_Temperatures, m_Times, m_heater_state = get_data('Growth_Curve_Stats.xlsx', True)
    Times, T_water, measured_T,heater_state, ODReadings, A = run_sim(1, int(np.size(m_Times)))
    m_ODReadings = -m_ODReadings #data was recorded incorrectly, so have to negative
    fig, ax1 = plt.subplots()
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel("Temperatures (°C)", color = 'red')
    ax1.plot(Times, m_Temperatures, label = 'Measured Temperature', color = 'purple')
    ax1.plot(Times, measured_T, label = 'Simulated Temperaure', color = 'red')
    plt.ylim(25,33)
    ax1.legend()

    ax2 = ax1.twinx()
    
    ax2.plot(Times, m_ODReadings, label = 'Measured ODReadings', color = 'black')
    ax2.plot(Times, ODReadings, label = 'Simulated ODReadings', color = 'blue')
    ax2.set_ylabel('Optical Denisty Readings', color = 'blue')
    plt.title("OD Readings & Temperature over time")
    ax2.legend(loc="lower right")
    plt.ylim(0,2)
    #plt.xlim(5000,15000)
    plt.grid()
    plt.show()

def compare_plots_chemostat():
    m_ODReadings, m_Temperatures, m_Times, m_heater_state = get_data('Chemostat_Data.xlsx')
    Times, T_water, measured_T,heater_state, ODReadings, A = run_sim(1, int(np.size(m_Times)), OD_control_type = 'Open-Loop')
    fig, ax1 = plt.subplots()
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel("Temperatures (°C)", color = 'red')
    ax1.plot(Times, m_Temperatures, label = 'Measured Temperature', color = 'purple')
    ax1.plot(Times, measured_T, label = 'Simulated Temperaure', color = 'red')
    plt.ylim(25,33)
    ax1.legend(loc='lower left')

    ax2 = ax1.twinx()
    
    ax2.plot(Times, m_ODReadings, label = 'Measured ODReadings', color = 'black')
    ax2.plot(Times, ODReadings, label = 'Simulated ODReadings', color = 'blue')
    ax2.set_ylabel('Optical Denisty Readings', color = 'blue')
    plt.title("OD Readings & Temperature over time")
    ax2.legend(loc="lower right")
    ##    plt.ylim(0,2)
    plt.xlim(13000,19000)   #Chemostat Data only valid  for x<20000
    plt.grid()
    plt.show()


# Run Simuation

def run_sim(dt, t_end, temp_control_type = 'Bang-Bang', OD_control_type = 'None'):

    n = int(t_end/dt)
    
    # Initialization of Arrays
    Times = np.linspace(0,t_end, n) #In seconds
    T_glass = np.ones(n) * T_ambient   # Glass vessel temperature (°C)
    T_water = np.ones(n) * T_initial    # Water temperature (°C)
    measured_T = np.copy(T_water) # Measured Temerature(Added noise)
    heater_state = np.zeros(n)  # Record of heater power over time
    A = np.ones(n) * A_initial # Record of bacteria population over time (Proportional to OD)
    OD = np.ones(n) *A_initial*C  #Optical Denisity
    heater_on_time = 0

    
    for k in range(1, n):

        if heater_state[k-1] == 1.0:
            heater_on_time+=1*dt
        else:
            heater_on_time = 0

        heater_delay = 139 # Heater power increases the longer heater remains on
        
        Q_heater = heater_state[k-1-heater_delay] * Q_heater_const 
        
    
        #Glass energy balance

        Q_to_water = h_gw * A_water_contact * (T_glass[k-1] - T_water[k-1])
        Q_loss_env = h_loss * A_env * (T_glass[k-1] - T_ambient)
        dT_glass =  (Q_heater - Q_to_water - Q_loss_env)  * dt / C_glass
    
        T_glass[k] = T_glass[k-1] + dT_glass + random.gauss(0, T_glass_sigma)*np.sqrt(dt)

        #Water energy balance

        # Water temperature changes according to heat received from the glass.
        dT_water = (Q_to_water - Q_loss_env) * dt / C_water
        T_water[k] = T_water[k-1] + dT_water 

        #Measurement noise for water temperature readings
        
        measured_T[k] = T_water[k]+random.gauss(0,T_measurement_sigma)
        
    

        #Update Bacteria Population (Due to bacteria growth)

        dt_A = A[k-1] * alpha_max * (1-A[k-1]/B)
        A[k] = A[k-1] + dt_A + random.gauss(0, Growth_Rate_sigma_mu/A[k-1])*np.sqrt(dt)

        #Update OD Measurement
    
        OD[k] = C*A[k] + random.gauss(0, OD_measurement_sigma)


        #Update Heater State due to temperature control

        heater_state[k] = temperature_control(measured_T, heater_state, k, dt, temp_control_type)

        #Update Temperature/OD Due to parameter readings

        flowrate = OD_control(OD,k, dt, OD_control_type)


        T_water[k] = T_water[k]*(1-flowrate/V_water)*dt+T_ambient*flowrate/V_water*dt

        A[k] = A[k]*(1-flowrate/V_water)*dt

        

        
    return Times, T_water, measured_T,heater_state, OD, A

compare_plots_chemostat()




    

